from __future__ import annotations

import hashlib
import zipfile
from pathlib import Path

import pytest

from ailabel_helpers import build_docx, build_pptx, build_xlsx, sign_package
from ops.pipelines.ai_label_removal import (
    AiLabelRemovalContext,
    remove_ai_labels_from_document,
)
from ops.pipelines.ai_label_removal.validators import extract_visible_text


def _run(tmp_path, src, name):
    ctx = AiLabelRemovalContext(request_id=f"t_{name}", source="test", original_filename=src.name)
    return remove_ai_labels_from_document(src, tmp_path / "out", ctx, audit_root=tmp_path / "audit")


def _sha(p):
    return hashlib.sha256(Path(p).read_bytes()).hexdigest()


class TestDocx:
    def test_removes_ai_and_preserves_visible_text(self, tmp_path):
        src = build_docx(tmp_path / "report.docx")
        before_text = extract_visible_text(src, "docx")
        before_hash = _sha(src)

        res = _run(tmp_path, src, "docx")

        assert res.status == "SUCCESS"
        assert res.output_path and res.output_path.exists()
        # Visible text preserved.
        assert extract_visible_text(res.output_path, "docx") == before_text
        # Original file untouched.
        assert _sha(src) == before_hash
        # AI markers removed; normal property kept.
        with zipfile.ZipFile(res.output_path) as z:
            custom = z.read("docProps/custom.xml").decode()
            core = z.read("docProps/core.xml").decode()
        assert "AIGenerated" not in custom and "Claude" not in custom
        assert "GrammarlyDocumentId" not in custom
        assert "Department" in custom  # normal metadata preserved
        assert "ChatGPT" not in core and "OpenAI" not in core
        assert "finance" in core.lower()  # non-AI keyword kept
        assert res.removed_labels

    def test_blocks_signed_package(self, tmp_path):
        src = build_docx(tmp_path / "report.docx")
        signed = sign_package(src)
        res = _run(tmp_path, signed, "docx_signed")
        assert res.status == "BLOCKED"
        assert res.output_path is None


class TestPptx:
    def test_removes_ai_and_preserves_slides(self, tmp_path):
        src = build_pptx(tmp_path / "deck.pptx")
        before_text = extract_visible_text(src, "pptx")
        before_hash = _sha(src)

        res = _run(tmp_path, src, "pptx")
        assert res.status == "SUCCESS"
        assert extract_visible_text(res.output_path, "pptx") == before_text
        assert _sha(src) == before_hash
        from pptx import Presentation
        assert len(Presentation(str(res.output_path)).slides) == 2
        with zipfile.ZipFile(res.output_path) as z:
            custom = z.read("docProps/custom.xml").decode()
        assert "AIGenerated" not in custom
        assert "GrammarlyDocumentId" not in custom
        assert "Owner" in custom


class TestXlsx:
    def test_removes_ai_and_preserves_sheets_and_formula(self, tmp_path):
        src = build_xlsx(tmp_path / "book.xlsx")
        before_hash = _sha(src)

        res = _run(tmp_path, src, "xlsx")
        assert res.status == "SUCCESS"
        assert _sha(src) == before_hash

        from openpyxl import load_workbook
        wb = load_workbook(str(res.output_path))
        assert wb.sheetnames == ["Data", "Notes"]
        ws = wb["Data"]
        assert ws["B2"].value == 10
        assert ws["B4"].value == "=SUM(B2:B3)"  # formula preserved
        assert wb["Notes"]["A1"].value == "Keep this sheet"
        with zipfile.ZipFile(res.output_path) as z:
            custom = z.read("docProps/custom.xml").decode()
        assert "AIGenerated" not in custom
        assert "GrammarlyDocumentId" not in custom
        assert "Region" in custom
