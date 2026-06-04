"""
axi_doctuning.py
────────────────
Doctuning / docfill pipeline utilities for the Axi bot.
Extracted from axi_bot.py (Phase C Task 2 refactor).

Provides the standalone (non-Telegram) doctuning utilities:
  _build_validate_system, _detect_template_reference_roles
  _doctuning_ws, _save_doctuning_memo, _find_existing_training_pair
  _save_doctuning_failure_case, _queue_doctuning_repair_case
  _request_repairman_investigation_for_doctuning_failure
  _read_document_text
  _validate_and_extract_context_sync, _nim_validate_document_sync
  _local_generate_fill_sync
  _check_ollama_alive_sync, _check_vram_free_gb_sync
  _repairman_fix_local_model_sync, _repairman_final_report_sync
  _local_fill_with_repair  (async)
  _xlsx_safe_text, _create_doctuning_xlsx_candidate, _xlsx_write_validation_failure
  _docfill_save_dpo_pair, _docfill_save_training_pair
  _doc_creator_120b_save, _doc_creator_120b_save_failed
  _docfill_save_to_omi

Note: async Telegram-handler functions (cmd_doctuning, _handle_docfill_file,
_approve_doctuning_for_chat, _run_doctuning_continue_for_chat, etc.) remain in
axi_bot.py because they depend on bot state (_PENDING_DOCFILL, etc.).
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
from datetime import datetime, timezone
from pathlib import Path

log = logging.getLogger("axi")

# ── Config (read from env at import time) ──────────────────────────────────────

AIMS_DOCUMENT_CREATOR_MODEL: str = (
    os.environ.get("AIMS_DOCUMENT_CREATOR_MODEL")
    or os.environ.get("AIMS_LOCAL_DRAFT_MODEL")
    or "axi_omi_sphere"
)
AIMS_DOCUMENT_TEACHER_MODEL: str = (
    os.environ.get("AIMS_DOCUMENT_TEACHER_MODEL")
    or os.environ.get("OMNIROUTE_TEACHER_MODEL")
    or "doc-training-standards-best"
)
AIMS_DOCUMENT_AUDIT_MODEL: str = (
    os.environ.get("AIMS_DOCUMENT_AUDIT_MODEL")
    or os.environ.get("OMNIROUTE_AUDIT_MODEL")
    or "doc-training-pair-audit-combo"
)
AIMS_DOC_CREATOR_TRAINING_DIR: Path = Path(
    os.environ.get("AIMS_DOC_CREATOR_TRAINING_DIR", "aims_workspace/training/document_creator_120b")
)

_DOCFILL_NIM_BASE = "https://integrate.api.nvidia.com/v1"
_DOCFILL_NIM_MODEL = "deepseek-ai/deepseek-v3-0324"

_DOCFILL_FILL_SYSTEM = (
    "You are an expert industrial document specialist. "
    "You receive a blank technical form and must fill in all fields with realistic values "
    "for an oil & gas / industrial maintenance context, following ISO 55001 and API standards. "
    "Output only the filled form text. No explanations, no comments."
)

_REPAIR_CMD_WHITELIST = (
    "ollama list",
    "ollama pull ",
    "ollama stop ",
    "ollama run ",
    "nvidia-smi",
    "curl http://localhost:11434",
    "systemctl status ollama",
    "systemctl restart ollama",
)

_REPAIRMAN_SYSTEM_DOCFILL = (
    "You are the AIMS Repairman — an expert repair agent for the AxiOMSphere platform.\n"
    "Always respond with a single JSON object:\n"
    '{"root_cause": "<one clear paragraph>", "files_changed": [], "patch_diff": "none", '
    '"tests_run": ["<safe shell command>", ...], "test_result": "not_run", '
    '"risk_level": "low", "rollback_notes": "<how to revert>"}\n'
    "Only include safe, non-destructive shell commands in tests_run. "
    "Focus on ollama model recovery: check if the model is loaded, VRAM state, "
    "whether the ollama service is running. Do not delete files or restart services "
    "unless risk_level is low."
)


# ── Validation system prompt builder ──────────────────────────────────────────

def _build_validate_system(standards: list[str]) -> str:
    std_line = (
        ", ".join(standards)
        if standards
        else "applicable industrial maintenance and safety standards"
    )
    return (
        "You are an expert industrial document quality auditor for the AIMS platform. "
        f"Check compliance against these standards: {std_line}. "
        "Assess whether the filled document is factually correct and free of hallucinations. "
        "Respond ONLY with valid JSON — no markdown fences, no extra text:\n"
        '{"hallucination_score": <0.0–1.0>, '
        '"issues": ["<specific issue>" ...], '
        '"standards_met": ["<standard>" ...], '
        '"verdict": "accept" | "reject", '
        '"reason": "<one sentence>"}'
    )


# ── Template/reference role detection ─────────────────────────────────────────

def _detect_template_reference_roles(doc_texts: list[dict]) -> tuple[dict, dict, str]:
    """Return (blank_doc, filled_doc, reason) using filename hints then content density."""
    TEMPLATE_HINTS = {"template", "blank", "form", "empty"}
    REFERENCE_HINTS = {"completed", "filled", "reference", "example", "gemini", "chatgpt"}

    def _score(name: str) -> tuple[int, int]:
        n = name.lower()
        return (
            sum(1 for h in TEMPLATE_HINTS if h in n),
            sum(1 for h in REFERENCE_HINTS if h in n),
        )

    if len(doc_texts) >= 2:
        t0, r0 = _score(doc_texts[0].get("name", ""))
        t1, r1 = _score(doc_texts[1].get("name", ""))

        if (r0 > t0) and (t1 > r1):
            return (
                doc_texts[1], doc_texts[0],
                f"filename: '{doc_texts[1]['name']}' → template, '{doc_texts[0]['name']}' → reference",
            )
        if (t0 > r0) and (r1 > t1):
            return (
                doc_texts[0], doc_texts[1],
                f"filename: '{doc_texts[0]['name']}' → template, '{doc_texts[1]['name']}' → reference",
            )
        if t0 > r0:
            return doc_texts[0], doc_texts[1], f"filename hint: '{doc_texts[0]['name']}' has template keyword"
        if t1 > r1:
            return doc_texts[1], doc_texts[0], f"filename hint: '{doc_texts[1]['name']}' has template keyword"
        if r0 > t0:
            return doc_texts[1], doc_texts[0], f"filename hint: '{doc_texts[0]['name']}' has reference keyword, treating doc[1] as template"
        if r1 > t1:
            return doc_texts[0], doc_texts[1], f"filename hint: '{doc_texts[1]['name']}' has reference keyword, treating doc[0] as template"

        _SPARSE = {"_", "__", "___", "N/A", "TBD", "[", "]", "...", "—"}

        def _density(text: str) -> float:
            toks = text.split()
            if not toks:
                return 0.0
            return sum(1 for t in toks if t not in _SPARSE) / len(toks)

        d0 = _density(doc_texts[0].get("text", ""))
        d1 = _density(doc_texts[1].get("text", ""))
        if d0 <= d1:
            return doc_texts[0], doc_texts[1], f"content density: '{doc_texts[0]['name']}' is sparser (likely blank)"
        return doc_texts[1], doc_texts[0], f"content density: '{doc_texts[1]['name']}' is sparser (likely blank)"

    if doc_texts:
        return doc_texts[0], doc_texts[0], "only one document provided"
    return {}, {}, "no documents"


# ── Workspace resolution ───────────────────────────────────────────────────────

def _doctuning_ws() -> Path:
    """Resolve the AIMS workspace root for doctuning artifacts."""
    for _env in ("AXI_DATA_ROOT", "AIMS_WORKSPACE"):
        _v = os.environ.get(_env)
        if _v:
            return Path(_v)
    for _candidate in (Path("/data"), Path("/aims_workspace")):
        if _candidate.exists():
            return _candidate
    return Path(__file__).resolve().parent.parent / "aims_workspace"


def _save_doctuning_memo(context: dict, blank_name: str, content_hash: str = None) -> Path:
    ws = _doctuning_ws()
    memo_dir = ws / "training" / "doctuning_memos"
    memo_dir.mkdir(parents=True, exist_ok=True)
    memo_path = memo_dir / f"doctuning_memo_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
    memo_data = {
        "blank_name": blank_name,
        "content_hash": content_hash,
        "created_at": datetime.now(timezone.utc).isoformat(),
        **context
    }
    memo_path.write_text(
        json.dumps(memo_data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return memo_path


def _find_existing_training_pair(content_hash: str) -> dict | None:
    """Check if training pair with this content hash already exists."""
    ws = _doctuning_ws()
    memo_dir = ws / "training" / "doctuning_memos"
    if not memo_dir.exists():
        return None

    for memo_file in memo_dir.glob("doctuning_memo_*.json"):
        try:
            memo_data = json.loads(memo_file.read_text(encoding="utf-8"))
            if memo_data.get("content_hash") == content_hash:
                return {
                    "memo_file": memo_file.name,
                    "created": memo_data.get("created_at", "unknown"),
                    "local_score": memo_data.get("local_model_score", "не оценено"),
                    "blank_name": memo_data.get("blank_name", "unknown"),
                }
        except Exception as e:
            log.warning("Failed to read memo %s: %s", memo_file.name, e)
            continue

    return None


# ── Failure case management ────────────────────────────────────────────────────

def _save_doctuning_failure_case(
    *,
    stage: str,
    reason: str,
    chat_id: int,
    state: dict | None = None,
    pair_path: str | None = None,
    severity: str = "warning",
    recoverable: bool = True,
) -> Path:
    """Save a doctuning failure case JSON for Repairman investigation."""
    import hashlib as _hashlib
    ws = _doctuning_ws()
    fail_dir = ws / "training" / "doctuning_failures"
    fail_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc)
    fname = f"{ts.strftime('%Y%m%d_%H%M%S')}_{stage}_{chat_id}.json"
    fail_path = fail_dir / fname

    _s = state or {}
    _source_text = str(_s.get("blank_text", "") or "")
    _source_hash = _hashlib.sha256(_source_text.encode()).hexdigest()[:16] if _source_text else ""

    data = {
        "source": "doctuning_batch_training_pair",
        "stage": stage,
        "reason": reason,
        "severity": severity,
        "recoverable": recoverable,
        "chat_id": chat_id,
        "pair_path": pair_path or "",
        "master_document_status": str(_s.get("master_document_status", "")),
        "master_document_error": str(_s.get("master_document_error", "")),
        "repairman_status": "pending",
        "created_at": ts.isoformat(),
        "state_summary": {
            "blank_name": str(_s.get("blank_name", "")),
            "filled_name": str(_s.get("filled_name", "")),
            "preferred_output_format": str(_s.get("preferred_output_format", "docx")),
            "source_hash": _source_hash,
            "task_id": str(_s.get("task_id", "")),
        },
    }
    try:
        fail_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        log.warning(
            "doctuning failure case saved: %s stage=%s reason=%s",
            fail_path, stage, reason[:120],
        )
    except Exception as _fe:
        log.error("doctuning failure case write failed: %s", _fe)
    return fail_path


def _queue_doctuning_repair_case(failure_path: Path, reason: str) -> None:
    """Append a repair queue entry to doctuning_repair_queue.jsonl."""
    ws = _doctuning_ws()
    queue_path = ws / "training" / "doctuning_repair_queue.jsonl"
    queue_path.parent.mkdir(parents=True, exist_ok=True)
    record = {
        "failure_path": str(failure_path),
        "reason": reason,
        "queued_at": datetime.now(timezone.utc).isoformat(),
        "status": "pending_repairman",
    }
    try:
        with queue_path.open("a", encoding="utf-8") as _qf:
            _qf.write(json.dumps(record, ensure_ascii=False) + "\n")
        log.warning("doctuning repair case queued: %s", failure_path)
    except Exception as _qe:
        log.error("doctuning repair queue write failed: %s", _qe)


def _request_repairman_investigation_for_doctuning_failure(failure_path: Path) -> bool:
    """Submit a doctuning failure case to RepairmanAPI (port 8010).

    Returns True if submitted, False if unavailable/failed (case queued instead).
    """
    _repairman_url = os.environ.get("AIMS_REPAIRMAN_API_URL", "http://localhost:8010")
    _new_status = "pending"
    try:
        _data = json.loads(failure_path.read_text(encoding="utf-8"))
        _task = (
            f"Doctuning failure: stage={_data.get('stage')} "
            f"reason={_data.get('reason', '')[:200]} "
            f"pair_path={_data.get('pair_path', '')} "
            f"severity={_data.get('severity')}"
        )
        import httpx as _httpx_rm
        with _httpx_rm.Client(timeout=5.0) as _c:
            _resp = _c.post(
                f"{_repairman_url}/repair",
                json={"task": _task, "mode": "inspect", "source": "doctuning_failure"},
            )
        _resp.raise_for_status()
        _new_status = "sent"
        _data["repairman_status"] = _new_status
        failure_path.write_text(json.dumps(_data, indent=2, ensure_ascii=False), encoding="utf-8")
        log.info("doctuning repairman investigation submitted: %s", failure_path)
        return True
    except Exception as _re:
        _err_str = str(_re).lower()
        _new_status = (
            "unavailable"
            if ("connection refused" in _err_str or "connectionrefused" in _err_str)
            else "failed"
        )
        try:
            _data = json.loads(failure_path.read_text(encoding="utf-8"))
            _data["repairman_status"] = _new_status
            failure_path.write_text(json.dumps(_data, indent=2, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass
        log.warning(
            "doctuning: repairman investigation not submitted (%s): %s", _new_status, _re
        )
        _queue_doctuning_repair_case(failure_path, str(_re)[:200])
        return False


# ── Document text extraction ───────────────────────────────────────────────────

def _read_document_text(file_path: Path) -> str:
    """Read text from .txt, .json, .docx, or .xlsx file."""
    suffix = file_path.suffix.lower()

    if suffix in (".txt", ".json"):
        return file_path.read_text(encoding="utf-8", errors="replace")

    elif suffix == ".docx":
        try:
            from docx import Document
            doc = Document(str(file_path))
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception as e:
            log.warning("docx read error: %s", e)
            return file_path.read_text(encoding="utf-8", errors="replace")

    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    line = " | ".join(str(v) if v is not None else "" for v in row)
                    if line.strip():
                        lines.append(line)
            return "\n".join(lines)
        except Exception as e:
            log.warning("xlsx read error: %s", e)
            return ""

    else:
        return file_path.read_text(encoding="utf-8", errors="replace")


# ── Validation pipeline ────────────────────────────────────────────────────────

def _validate_and_extract_context_sync(blank_text: str, filled_text: str) -> dict:
    """Combined: extract context + validate + produce revised reference via OmniRoute.

    Returns dict with keys: context, validation, revised_reference, error (on failure).
    """
    from axi_omniroute_session import _call_doctuning_openai_omniroute_sync, _extract_omniroute_content_from_body

    model = (
        os.environ.get("AIMS_DOCTUNING_VALIDATE_MODEL")
        or os.environ.get("OMNIROUTE_AUDIT_MODEL")
        or os.environ.get("AIMS_DOCTUNING_MODEL")
        or "doc-training-pair-audit-combo"
    )

    def _failed_validation(reason: str) -> dict:
        return {
            "status": "failed",
            "hallucination_score": None,
            "evidence_risk": None,
            "issues": ["teacher_audit_failed"],
            "standards_met": [],
            "reason": reason,
            "missing_fields": [],
            "unsupported_claims": [],
            "validation_gaps": [],
            "recommendations": [],
        }

    combined_system = (
        "You are an expert industrial document specialist and quality auditor for the AIMS platform. "
        "Perform THREE tasks and respond with a single valid JSON object. "
        "No markdown fences, no extra text outside the JSON.\n\n"
        "Task 1 — CLASSIFY and extract context:\n"
        '  "context": {\n'
        '    "doc_type": "<procedure|specification|report|form|checklist|...>",\n'
        '    "equipment_type": "<equipment or process name>",\n'
        '    "industry": "<sector>",\n'
        '    "key_terms": ["<term>", ...],\n'
        '    "standards": ["<applicable standard>", ...]\n'
        "  }\n\n"
        "Task 2 — VALIDATE the filled reference for hallucinations, fabricated values, "
        "missing required fields, and standards compliance (ISO, API, ASME, OSHA, IEC, NFPA, "
        "IEEE, EN, ANSI, PAS, BS). Do NOT invent page or section references.\n"
        "  Status rules — use exactly one of these values:\n"
        '    "failed"  — any of: hallucination_score ≥ 0.5; fabricated standard codes or\n'
        "               equipment specifications; required normative sections completely absent.\n"
        '    "warning" — issues present but none are fatal; document is usable with corrections.\n'
        "               ALWAYS use 'warning' (not 'passed') when admin/template fields remain\n"
        "               unfilled: TBD, TODO, Placeholder, [Date], [Author], [Dept], ???.\n"
        '    "passed"  — only when ALL of the following hold: no required fields missing,\n'
        "               hallucination_score < 0.2, no fabricated data, no TBD/placeholder\n"
        "               admin fields, no unsupported normative claims.\n"
        '  "validation": {\n'
        '    "status": "passed" | "warning" | "failed",\n'
        '    "hallucination_score": <0.0–1.0, fabrications only>,\n'
        '    "evidence_risk": <0.0–1.0, unsubstantiated assertions for this doc type>,\n'
        '    "issues": ["<specific issue>", ...],\n'
        '    "standards_met": ["<standard>", ...],\n'
        '    "reason": "<one sentence overall verdict>",\n'
        '    "missing_fields": ["<field or section missing>", ...],\n'
        '    "unsupported_claims": ["<normative claim with no traceable evidence>", ...],\n'
        '    "validation_gaps": ["<check not performed due to doc type or missing data>", ...],\n'
        '    "recommendations": ["<actionable fix>", ...]\n'
        "  }\n\n"
        "Task 3 — PRODUCE a revised reference document:\n"
        '  "revised_reference": "<full revised document text>"\n'
        "  Rules:\n"
        "  - Do NOT create a new document from scratch.\n"
        "  - Start from the supplied filled reference.\n"
        "  - Preserve all valid content and structure.\n"
        "  - Apply only corrections for issues found in Task 2.\n"
        "  - If no changes are needed, return the supplied reference unchanged.\n"
        "  - Always return revised_reference as a non-empty string."
    )

    prompt = (
        f"BLANK TEMPLATE:\n{blank_text[:3000]}\n\n"
        f"FILLED REFERENCE (to audit and revise):\n{filled_text[:4000]}\n\n"
        "Perform all three tasks and return a single JSON object. "
        "Return ONLY raw JSON. Do not use markdown fences. Do not include text outside JSON."
    )

    log.info(
        "doctuning validate+context+revise: model=%s blank_len=%d filled_len=%d",
        model, len(blank_text), len(filled_text),
    )

    status, raw = _call_doctuning_openai_omniroute_sync(
        model=model,
        system=combined_system,
        prompt=prompt,
        max_tokens=8192,
        temperature=0.1,
        timeout=300.0,
    )

    log.info("doctuning validate+context+revise: status=%d raw_len=%d", status, len(raw))

    if status != 200:
        err_key = f"omniroute_status_{status}"
        return {
            "context": {},
            "validation": _failed_validation(f"{err_key}: {raw[:200]}"),
            "revised_reference": None,
            "error": err_key,
            "raw": raw[:500],
        }

    if not raw:
        return {
            "context": {},
            "validation": _failed_validation("teacher_audit_empty_response"),
            "revised_reference": None,
            "error": "teacher_audit_empty_response",
            "raw": "",
        }

    if raw.lstrip().startswith("data:") and "chat.completion.chunk" in raw:
        raw = _extract_omniroute_content_from_body(raw)

    cleaned = raw.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.lstrip("`")
        if cleaned.startswith("json"):
            cleaned = cleaned[4:]
        cleaned = cleaned.strip()
    if cleaned.endswith("```"):
        cleaned = cleaned[:-3].strip()
    cleaned = cleaned.strip("`").strip()
    if cleaned.lower().startswith("json"):
        cleaned = cleaned[4:].strip()

    try:
        result = json.loads(cleaned)
    except Exception as exc:
        return {
            "context": {},
            "validation": _failed_validation(f"json_parse_failed: {exc}; raw_start={cleaned[:500]}"),
            "revised_reference": None,
            "error": "json_parse_failed",
            "raw": cleaned[:500],
        }

    if "context" not in result:
        result["context"] = {}
    if "validation" not in result:
        result["validation"] = _failed_validation("no validation returned")
    else:
        v = result["validation"]
        if "status" not in v:
            v["status"] = "passed" if not v.get("issues") else "warning"
        for field in ("missing_fields", "unsupported_claims", "recommendations", "validation_gaps"):
            if field not in v:
                v[field] = []
        if "evidence_risk" not in v:
            v["evidence_risk"] = None

        import re as _re
        _TBD_PAT = _re.compile(
            r"\b(TBD|TODO|PLACEHOLDER|TO BE DETERMINED|TO BE CONFIRMED)\b"
            r"|\[Date\]|\[DATE\]|\[Author\]|\[AUTHOR\]|\[Dept\]|\[DEPT\]"
            r"|\[Department\]|\[Reviewer\]|\[REVIEWER\]|\[Approver\]|\[APPROVER\]"
            r"|\[Name\]|\[Title\]|\[Division\]|\[Unit\]"
            r"|\?\?\?",
            _re.IGNORECASE,
        )
        if v.get("status") == "passed" and _TBD_PAT.search(filled_text):
            v["status"] = "warning"
            issues = v.setdefault("issues", [])
            if "admin_fields_incomplete" not in issues:
                issues.append("admin_fields_incomplete")
            mf = v.setdefault("missing_fields", [])
            _tbd_note = "Admin/template placeholder fields (TBD, [Date], [Author], etc.) not filled in"
            if _tbd_note not in mf:
                mf.append(_tbd_note)

    if not result.get("revised_reference"):
        result["revised_reference"] = filled_text
        issues = result["validation"].setdefault("issues", [])
        if "teacher_revision_missing" not in issues:
            issues.append("teacher_revision_missing")

    return result


def _nim_validate_document_sync(blank_text: str, filled_text: str, standards: list[str]) -> dict:
    """DEPRECATED: Use _validate_and_extract_context_sync instead. Kept for backward compat."""
    try:
        from openai import OpenAI as _OAI
    except ImportError:
        return {"verdict": "accept", "hallucination_score": 0.0, "issues": [], "reason": "openai not installed — skipped validation", "standards_met": []}

    api_key = os.environ.get("NVIDIA_API_KEY", "").strip()
    if not api_key:
        return {"verdict": "accept", "hallucination_score": 0.0, "issues": [], "reason": "NVIDIA_API_KEY not set — skipped validation", "standards_met": []}

    client = _OAI(api_key=api_key, base_url=_DOCFILL_NIM_BASE)
    prompt = (
        f"BLANK TEMPLATE:\n{blank_text[:3000]}\n\n"
        f"FILLED EXAMPLE (to audit):\n{filled_text[:4000]}\n\n"
        "Audit the filled example for hallucinations, fabricated values, and standards compliance. "
        "Return JSON as specified."
    )
    try:
        resp = client.chat.completions.create(
            model=_DOCFILL_NIM_MODEL,
            messages=[
                {"role": "system", "content": _build_validate_system(standards)},
                {"role": "user", "content": prompt},
            ],
            temperature=0.1,
            max_tokens=512,
        )
        raw = (resp.choices[0].message.content or "").strip()
        raw = raw.strip("`").strip()
        if raw.startswith("json"):
            raw = raw[4:].strip()
        return json.loads(raw)
    except Exception as e:
        log.warning("docfill nim validate error: %s", e)
        return {"verdict": "accept", "hallucination_score": 0.0, "issues": [], "reason": f"validation error ({e}) — accepted by default", "standards_met": []}


# ── Local Ollama generation ────────────────────────────────────────────────────

def _local_generate_fill_sync(blank_text: str) -> str | None:
    """Ask local ollama model to fill the blank form (DPO rejected — model's current output)."""
    try:
        import httpx
    except ImportError:
        log.warning("local fill: httpx not available")
        return None

    ollama_url = os.environ.get("OLLAMA_LOCAL_URL", "http://localhost:11434").rstrip("/")
    model = AIMS_DOCUMENT_CREATOR_MODEL

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": _DOCFILL_FILL_SYSTEM},
            {"role": "user", "content": f"Fill in this blank form:\n\n{blank_text[:4000]}"},
        ],
        "stream": False,
        "temperature": 0.3,
        "max_tokens": 2048,
    }

    try:
        with httpx.Client(timeout=120.0) as client:
            resp = client.post(
                f"{ollama_url}/v1/chat/completions",
                json=payload,
                headers={"Content-Type": "application/json"},
            )
        if resp.status_code != 200:
            log.warning("local fill: ollama returned status %d", resp.status_code)
            return None
        data = resp.json()
        result = (
            data.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
            or ""
        ).strip()
        return result if len(result) > 80 else None
    except Exception as e:
        log.warning("doctuning local generate error: %s", e)
        return None


def _check_ollama_alive_sync() -> bool:
    """Return True if the local Ollama service responds on /api/tags."""
    import urllib.request as _ur
    ollama_url = os.environ.get("OLLAMA_LOCAL_URL", "http://localhost:11434").rstrip("/")
    try:
        with _ur.urlopen(f"{ollama_url}/api/tags", timeout=5) as r:
            return r.status == 200
    except Exception:
        return False


def _check_vram_free_gb_sync() -> float:
    """Return free VRAM in GB on the first GPU via nvidia-smi, or 999.0 if unavailable."""
    import subprocess as _sp
    try:
        out = _sp.check_output(
            ["nvidia-smi", "--query-gpu=memory.free", "--format=csv,noheader,nounits"],
            timeout=10, text=True,
        )
        first_line = out.strip().splitlines()[0]
        if first_line != "[N/A]":
            mb = float(first_line)
            return mb / 1024.0

        out = _sp.check_output(
            ["nvidia-smi", "--query-gpu=utilization.gpu", "--format=csv,noheader,nounits"],
            timeout=10, text=True,
        )
        util_pct = float(out.strip().splitlines()[0])
        if util_pct >= 92.0:
            return 0.0
        else:
            return 999.0
    except Exception:
        return 999.0


# ── Repairman integration ──────────────────────────────────────────────────────

def _repairman_fix_local_model_sync(model: str) -> str:
    """Call repairman gateway to diagnose and recover the local model. Returns root_cause string."""
    import subprocess as _sp
    import json as _json

    gateway_url = "http://localhost:8082/v1"
    token = os.environ.get("AIMS_CLAUDE_PROXY_TOKEN", "aims-local-repair-token")
    problem = (
        f"The local Ollama model '{model}' failed to return a completion in the doctuning "
        "pipeline (returned None or raised an exception). Likely causes: model not loaded, "
        "VRAM exhaustion (DGX 128 GB — check if two large models are loaded simultaneously), "
        "ollama service not running, or connection refused on port 11434. "
        "Diagnose and provide safe recovery commands in 'tests_run'."
    )
    payload = {
        "model": "aims-repairman-nemotron",
        "messages": [
            {"role": "system", "content": _REPAIRMAN_SYSTEM_DOCFILL},
            {"role": "user",   "content": problem},
        ],
        "temperature": 0.1,
        "max_tokens": 512,
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    _default_root_cause = "Repairman unreachable — manual Ollama check required."

    try:
        try:
            from openai import OpenAI as _OAI
            client = _OAI(api_key=token, base_url=gateway_url)
            resp_oa = client.chat.completions.create(
                model="aims-repairman-nemotron",
                messages=payload["messages"],
                temperature=0.1,
                max_tokens=512,
            )
            raw = (resp_oa.choices[0].message.content or "").strip()
        except ImportError:
            import httpx as _httpx
            with _httpx.Client(timeout=90.0) as _c:
                _r = _c.post(f"{gateway_url}/chat/completions", json=payload, headers=headers)
            raw = (
                _r.json().get("choices", [{}])[0]
                .get("message", {})
                .get("content", "")
                or ""
            ).strip()

        if not raw:
            return _default_root_cause

        # Parse JSON response
        _raw_stripped = raw.strip().strip("`").strip()
        if _raw_stripped.startswith("json"):
            _raw_stripped = _raw_stripped[4:].strip()
        try:
            result = _json.loads(_raw_stripped)
        except Exception:
            return raw[:400]

        root_cause = result.get("root_cause", _default_root_cause)

        # Execute safe recovery commands (whitelist enforced)
        tests_run = result.get("tests_run", [])
        if result.get("risk_level") == "low" and isinstance(tests_run, list):
            for cmd in tests_run[:3]:
                if not isinstance(cmd, str):
                    continue
                if not any(cmd.startswith(ok) for ok in _REPAIR_CMD_WHITELIST):
                    log.warning("repairman: command not in whitelist, skipped: %r", cmd[:100])
                    continue
                try:
                    result_proc = _sp.run(
                        cmd.split(), capture_output=True, text=True, timeout=30
                    )
                    log.info("repairman executed: %r → rc=%d stdout=%s",
                             cmd, result_proc.returncode, result_proc.stdout[:200])
                except Exception as _ce:
                    log.warning("repairman command failed: %r: %s", cmd, _ce)

        return root_cause

    except Exception as e:
        log.warning("_repairman_fix_local_model_sync error: %s", e)
        return _default_root_cause


def _repairman_final_report_sync(model: str, cycles: int, last_root_cause: str) -> str:
    """After exhausting all repair cycles, ask repairman to write a human-readable failure report."""
    gateway_url = "http://localhost:8082/v1"
    token = os.environ.get("AIMS_CLAUDE_PROXY_TOKEN", "aims-local-repair-token")
    problem = (
        f"You attempted to repair the local Ollama model '{model}' {cycles} times. "
        f"Last diagnosed root cause: {last_root_cause}\n"
        f"All {cycles} repair cycles failed — generation still returns no valid output. "
        "Write a concise 2–3 sentence report for the engineer explaining:\n"
        "1. What the root cause appears to be.\n"
        "2. What specific manual action is required to resolve it.\n"
        "Plain text only, no JSON, no markdown."
    )
    system_msg = (
        "You are the AIMS Repairman reporting a repair failure to a human engineer. "
        "Be concise and actionable."
    )
    payload = {
        "model": "aims-repairman-nemotron",
        "messages": [
            {"role": "system", "content": system_msg},
            {"role": "user",   "content": problem},
        ],
        "temperature": 0.1,
        "max_tokens": 256,
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    try:
        try:
            from openai import OpenAI as _OAI
            client = _OAI(api_key=token, base_url=gateway_url)
            resp_oa = client.chat.completions.create(
                model="aims-repairman-nemotron",
                messages=payload["messages"],
                temperature=0.1,
                max_tokens=256,
            )
            return (resp_oa.choices[0].message.content or "").strip()
        except ImportError:
            import httpx as _httpx
            with _httpx.Client(timeout=60.0) as _c:
                _r = _c.post(f"{gateway_url}/chat/completions", json=payload, headers=headers)
            return (
                _r.json().get("choices", [{}])[0]
                .get("message", {})
                .get("content", "")
                or ""
            ).strip()
    except Exception as e:
        return (
            f"Repairman unreachable after {cycles} cycles. "
            f"Last known cause: {last_root_cause}. Manual intervention required."
        )


async def _local_fill_with_repair(
    blank_text: str, max_repair_cycles: int = 5
) -> tuple[str | None, str]:
    """Generate local fill. On first failure enters the repairman repair loop.

    Returns (filled_text, repairman_report).
    filled_text is None when all cycles failed; repairman_report contains the final diagnosis.
    """
    loop = asyncio.get_event_loop()
    model = AIMS_DOCUMENT_CREATOR_MODEL

    result = await loop.run_in_executor(None, _local_generate_fill_sync, blank_text)
    if result:
        return result, ""

    _gateway_url = "http://localhost:8082"
    try:
        import httpx as _httpx_hc
        with _httpx_hc.Client(timeout=5.0) as _hc:
            _hc.get(f"{_gateway_url}/health")
    except Exception as _hc_exc:
        _hc_msg = str(_hc_exc)
        log.warning("doctuning: repairman unreachable, skipping repair loop: %s", _hc_msg)
        return None, f"repairman unavailable: {_hc_msg}"

    log.warning(
        "doctuning: initial generation failed — starting repairman loop (max %d cycles)",
        max_repair_cycles,
    )
    last_root_cause = "unknown"

    for cycle in range(1, max_repair_cycles + 1):
        log.info("doctuning: repairman cycle %d/%d", cycle, max_repair_cycles)
        last_root_cause = await loop.run_in_executor(
            None, _repairman_fix_local_model_sync, model
        )
        log.info("repairman cycle %d root_cause: %s", cycle, last_root_cause)
        await asyncio.sleep(20)
        alive = await loop.run_in_executor(None, _check_ollama_alive_sync)
        if not alive:
            log.warning("repairman cycle %d: ollama still down after repair", cycle)
            continue
        result = await loop.run_in_executor(None, _local_generate_fill_sync, blank_text)
        if result:
            log.info("doctuning: generation succeeded after repairman cycle %d", cycle)
            return result, ""
        log.warning("repairman cycle %d: service up but generation still fails", cycle)

    log.error("doctuning: repairman exhausted %d cycles, requesting final report", max_repair_cycles)
    final_report = await loop.run_in_executor(
        None, _repairman_final_report_sync, model, max_repair_cycles, last_root_cause
    )
    log.error("repairman final report: %s", final_report)
    return None, final_report


# ── XLSX candidate builder ─────────────────────────────────────────────────────

def _xlsx_safe_text(value: object) -> str:
    """Sanitize a value for an openpyxl string cell — prevent Excel formula injection."""
    s = str(value) if value is not None else ""
    if s and s[0] in ("=", "+", "-", "@", "\t", "\r", "|"):
        s = "'" + s
    return s


def _create_doctuning_xlsx_candidate(
    template_path: "Path | None",
    reference_path: "Path | None",
    chosen_text: str,
    context: dict,
    validation: dict,
    output_path: "Path",
) -> "Path":
    """Build a validated XLSX candidate using the incoming workbook as base."""
    import shutil
    import subprocess
    import tempfile

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as _ie:
        raise RuntimeError("openpyxl not available") from _ie

    source_path: "Path | None" = None
    for candidate_src in (reference_path, template_path):
        if candidate_src and Path(candidate_src).exists() and Path(candidate_src).suffix.lower() == ".xlsx":
            source_path = Path(candidate_src)
            break

    if source_path is None:
        raise FileNotFoundError(
            "No source .xlsx workbook available "
            "(reference_path and template_path both absent or not .xlsx)"
        )

    try:
        shutil.copy2(str(source_path), str(output_path))
        log.info("doctuning xlsx: copied %s → %s", source_path, output_path)
    except Exception as _e:
        raise RuntimeError(f"Could not copy source workbook: {_e}") from _e

    try:
        wb = load_workbook(str(output_path))
    except Exception as _e:
        Path(output_path).unlink(missing_ok=True)
        raise RuntimeError(f"Could not open copied workbook: {_e}") from _e

    if "AIMS_Review" in wb.sheetnames:
        del wb["AIMS_Review"]
    ws = wb.create_sheet("AIMS_Review")

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    wrap = Alignment(wrap_text=True, vertical="top")

    def _row(label: str, value: str, row_idx: int) -> None:
        ws.cell(row=row_idx, column=1, value=label).font = header_font
        ws.cell(row=row_idx, column=1).fill = header_fill
        cell = ws.cell(row=row_idx, column=2, value=_xlsx_safe_text(value))
        cell.alignment = wrap

    row = 1
    ws.cell(row=row, column=1, value="AIMS Doctuning Review").font = Font(bold=True, size=13)
    row += 1

    _row("Document type",       context.get("doc_type", "—"),                                   row); row += 1
    _row("Equipment type",      context.get("equipment_type", "—"),                              row); row += 1
    _row("Industry",            context.get("industry", "—"),                                   row); row += 1
    _row("Standards",           ", ".join(context.get("standards", [])) or "—",                  row); row += 1
    _row("Key terms",           ", ".join(context.get("key_terms", [])) or "—",                  row); row += 1
    _row("Validation status",   validation.get("status", "—"),                                   row); row += 1
    h = validation.get("hallucination_score")
    _row("Hallucination score", f"{h:.3f}" if h is not None else "—",                            row); row += 1
    er = validation.get("evidence_risk")
    _row("Evidence/substantiation risk", f"{er:.3f}" if isinstance(er, float) else (str(er) if er is not None else "—"), row); row += 1
    _row("Issues",              "\n".join(validation.get("issues", [])) or "none",               row); row += 1
    _row("Missing fields",      "\n".join(validation.get("missing_fields", [])) or "none",      row); row += 1
    _row("Validation gaps",     "\n".join(validation.get("validation_gaps", [])) or "none",     row); row += 1
    _row("Unsupported claims",  "\n".join(validation.get("unsupported_claims", [])) or "none",  row); row += 1
    _row("Recommendations",     "\n".join(validation.get("recommendations", [])) or "none",     row); row += 1
    _row("Reason",              validation.get("reason", "—"),                                   row); row += 1

    row += 1
    ws.cell(row=row, column=1, value="Revised reference text").font = Font(bold=True)
    row += 1
    cell = ws.cell(row=row, column=2, value=_xlsx_safe_text(chosen_text))
    cell.alignment = wrap

    row += 2
    _row("Note", "Original workbook preserved; AIMS review added as separate sheet.", row)

    ws.column_dimensions[get_column_letter(1)].width = 26
    ws.column_dimensions[get_column_letter(2)].width = 90

    try:
        wb.save(str(output_path))
        log.info("doctuning xlsx: saved to %s", output_path)
    except Exception as _e:
        Path(output_path).unlink(missing_ok=True)
        raise RuntimeError(f"openpyxl save failed: {_e}") from _e

    # Validation 1: openpyxl reopen
    try:
        load_workbook(str(output_path), data_only=False)
        log.info("doctuning xlsx: openpyxl validation passed")
    except Exception as exc:
        log.warning("doctuning xlsx validation failed: %s", exc)
        _xlsx_write_validation_failure(output_path, source_path, exc)
        Path(output_path).unlink(missing_ok=True)
        raise ValueError(f"XLSX failed openpyxl validation: {exc}") from exc

    # Validation 2: LibreOffice headless convert (if available)
    lo_bin = shutil.which("libreoffice") or shutil.which("soffice")
    if lo_bin:
        try:
            with tempfile.TemporaryDirectory() as _tmpdir:
                result = subprocess.run(
                    [lo_bin, "--headless", "--convert-to", "xlsx",
                     "--outdir", _tmpdir, str(output_path)],
                    capture_output=True, text=True, timeout=90,
                )
            if result.returncode != 0:
                raise RuntimeError(f"exit {result.returncode}: {result.stderr[:400]}")
            log.info("doctuning xlsx: LibreOffice validation passed")
        except Exception as exc:
            log.warning("doctuning xlsx libreoffice validation failed: %s", exc)
            _xlsx_write_validation_failure(output_path, source_path, exc)
            Path(output_path).unlink(missing_ok=True)
            raise ValueError(f"XLSX failed LibreOffice validation: {exc}") from exc
    else:
        log.info("doctuning xlsx: libreoffice not found — skipping LO validation")

    return Path(output_path)


def _xlsx_write_validation_failure(output_path: "Path", source_path: "Path | None", exc: Exception) -> None:
    """Write a sidecar failure note to doctuning_debug/. Best-effort, not sent to user."""
    try:
        debug_dir = _doctuning_ws() / "training" / "doctuning_debug"
        debug_dir.mkdir(parents=True, exist_ok=True)
        sidecar = debug_dir / (Path(output_path).stem + "_validation_failure.txt")
        sidecar.write_text(
            f"XLSX validation failed\nsource: {source_path}\nerror: {exc}\n",
            encoding="utf-8",
        )
    except Exception:
        pass


# ── Training data savers ───────────────────────────────────────────────────────

def _docfill_save_dpo_pair(blank_text: str, chosen_text: str, rejected_text: str, doc_type: str) -> None:
    """Save doctuning DPO pair (human expert chosen, local model rejected)."""
    ws = Path(__file__).resolve().parent.parent
    dpo_path = ws / "data/training/standard_dpo_pairs.jsonl"
    dpo_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with dpo_path.open("a", encoding="utf-8") as f:
            f.write(json.dumps({
                "prompt": blank_text,
                "chosen": chosen_text,
                "chosen_score": 1.0,
                "rejected": rejected_text,
                "rejected_score": 0.5,
                "rejected_feedback": "Local model fill — to be improved toward human expert reference",
                "pipeline": f"doctuning_{doc_type}",
                "target_model": AIMS_DOCUMENT_CREATOR_MODEL,
                "target_skill": "document_structure_generation",
                "rejected_model": AIMS_DOCUMENT_CREATOR_MODEL,
                "teacher_model": AIMS_DOCUMENT_TEACHER_MODEL,
                "audit_model": AIMS_DOCUMENT_AUDIT_MODEL,
                "teacher_models": {
                    "standards_discovery": AIMS_DOCUMENT_TEACHER_MODEL,
                    "audit": AIMS_DOCUMENT_AUDIT_MODEL,
                    "quality_judge": AIMS_DOCUMENT_TEACHER_MODEL,
                },
            }, ensure_ascii=False) + "\n")
        log.info("doctuning: DPO pair saved to standard_dpo_pairs.jsonl")
    except Exception as e:
        log.warning("doctuning dpo pair save failed: %s", e)


def _docfill_save_training_pair(blank_text: str, filled_text: str, doc_type: str, source_name: str) -> Path:
    """Append a (blank → filled) training pair to the docfill dataset."""
    ws = Path(__file__).resolve().parent.parent
    out_dir = ws / "ops/ft/data/docfill_v1"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "train_docfill_v1.jsonl"

    system_msg = (
        "You are an AIMS document specialist. When given a blank technical form, "
        "fill in all fields accurately for the specified equipment type based on "
        "industrial maintenance and safety standards."
    )
    pair = {
        "messages": [
            {"role": "system", "content": system_msg},
            {"role": "user", "content": f"Fill in this form:\n\n{blank_text}"},
            {"role": "assistant", "content": filled_text},
        ],
        "_meta": {
            "source": "axi_docfill_upload",
            "doc_type": doc_type,
            "blank_name": source_name,
            "saved_at": datetime.now(timezone.utc).isoformat(),
            "target_model": AIMS_DOCUMENT_CREATOR_MODEL,
            "target_skill": "document_structure_generation",
            "rejected_model": AIMS_DOCUMENT_CREATOR_MODEL,
            "teacher_model": AIMS_DOCUMENT_TEACHER_MODEL,
            "audit_model": AIMS_DOCUMENT_AUDIT_MODEL,
            "teacher_models": {
                "standards_discovery": AIMS_DOCUMENT_TEACHER_MODEL,
                "audit": AIMS_DOCUMENT_AUDIT_MODEL,
                "quality_judge": AIMS_DOCUMENT_TEACHER_MODEL,
            },
        },
    }
    with out_file.open("a", encoding="utf-8") as f:
        f.write(json.dumps(pair, ensure_ascii=False) + "\n")
    return out_file


def _doc_creator_120b_save(
    *,
    pair_type: str,
    blank_text: str,
    filled_text: str,
    doc_type: str,
    source_name: str,
    extra_meta: dict | None = None,
) -> Path:
    """Save a training record to the 120B document-creator training pool."""
    subdir = AIMS_DOC_CREATOR_TRAINING_DIR / pair_type
    subdir.mkdir(parents=True, exist_ok=True)
    out_file = subdir / f"train_{pair_type}.jsonl"
    meta: dict = {
        "source": source_name,
        "doc_type": doc_type,
        "pair_type": pair_type,
        "saved_at": datetime.now(timezone.utc).isoformat(),
        "target_model": AIMS_DOCUMENT_CREATOR_MODEL,
        "target_skill": "document_structure_generation",
        "teacher_model": AIMS_DOCUMENT_TEACHER_MODEL,
        "audit_model": AIMS_DOCUMENT_AUDIT_MODEL,
    }
    if extra_meta:
        meta.update(extra_meta)
    record = {
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are an AIMS document specialist. When given a blank technical form, "
                    "fill in all fields accurately for the specified equipment type based on "
                    "industrial maintenance and safety standards."
                ),
            },
            {"role": "user", "content": f"Fill in this form:\n\n{blank_text}"},
            {"role": "assistant", "content": filled_text},
        ],
        "_meta": meta,
    }
    with out_file.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
    return out_file


def _doc_creator_120b_save_failed(
    *,
    blank_text: str,
    doc_type: str,
    source_name: str,
    repair_report: str,
) -> None:
    """Record a failed 120B generation attempt for later analysis."""
    subdir = AIMS_DOC_CREATOR_TRAINING_DIR / "failed_generation_cases"
    subdir.mkdir(parents=True, exist_ok=True)
    out_file = subdir / "failed_generation_cases.jsonl"
    record = {
        "blank_text": blank_text[:2000],
        "doc_type": doc_type,
        "source_name": source_name,
        "repair_report": repair_report,
        "saved_at": datetime.now(timezone.utc).isoformat(),
        "target_model": AIMS_DOCUMENT_CREATOR_MODEL,
    }
    try:
        with out_file.open("a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except Exception as e:
        log.warning("doc_creator_120b failed-case save error: %s", e)


def _docfill_save_to_omi(title: str, content: str, doc_type: str) -> str | None:
    """Optional future path: register master document via OmiAgent POST /documents.

    Only call this when a real /documents endpoint is confirmed to exist.
    Returns doc_id string on success. Raises RuntimeError on failure.
    """
    import uuid as _uuid
    import urllib.request as _ur
    import urllib.error as _ue

    base_url = (
        os.environ.get("OMI_API_URL")
        or os.environ.get("OMI_URL")
        or os.environ.get("AIMS_OMI_URL")
        or "http://localhost:8008"
    ).rstrip("/")
    url = f"{base_url}/documents"
    task_id = f"doctuning-{_uuid.uuid4().hex[:12]}"

    payload = json.dumps({
        "task_id": task_id,
        "title": title,
        "doc_type": doc_type,
        "status": "master",
        "quality_score": 0.0,
        "summary": (content or "")[:500],
        "metadata": {
            "source": "doctuning_batch_training_pair",
            "created_at": datetime.now(timezone.utc).isoformat(),
        },
    }).encode()

    log.info("docfill_save_to_omi: POST %s task_id=%s title=%r", url, task_id, title[:80])
    try:
        req = _ur.Request(
            url,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with _ur.urlopen(req, timeout=15) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            result = json.loads(body)
            doc_id = (
                result.get("doc_id")
                or result.get("id")
                or result.get("document_id")
            )
            if doc_id:
                log.info("docfill_save_to_omi: registered doc_id=%s", doc_id)
                return str(doc_id)
            raise RuntimeError(f"doc_id not found in response: {body[:200]}")
    except _ue.HTTPError as _he:
        raise RuntimeError(f"OmiAgent /documents HTTP {_he.code}: {_he.read()[:200]!r}") from _he
    except Exception as e:
        raise RuntimeError(f"OmiAgent /documents error: {e}") from e
